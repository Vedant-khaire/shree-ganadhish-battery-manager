from datetime import date, datetime
import datetime as dt_module
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from supabase import Client

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1E3A5F")  # dark navy
CENTER = Alignment(horizontal="center", vertical="center")

ZEBRA_FILL = PatternFill(fill_type="solid", fgColor="F9F9F9")
GREEN_FILL = PatternFill(fill_type="solid", fgColor="E2EFDA")  # soft green
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC") # soft yellow
RED_FILL = PatternFill(fill_type="solid", fgColor="FFD9D9")    # soft red
CARD_LABEL_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")
CARD_VALUE_FILL = PatternFill(fill_type="solid", fgColor="EAF2F8")

# ---------------------------------------------------------------------------
# Normalization & Helpers
# ---------------------------------------------------------------------------

def _normalize_payment_mode(mode: Optional[str]) -> str:
    if not mode:
        return "Other"
    mode_strip = mode.strip().lower()
    if not mode_strip or mode_strip in ("n/a", "null", "none"):
        return "Other"
    
    if mode_strip in ("cash", "c"):
        return "Cash"
    if mode_strip in ("upi", "gpay", "phonepe", "paytm", "g_pay", "phone_pe"):
        return "UPI"
    if mode_strip in ("net_banking", "netbanking", "net banking", "online", "rtgs", "neft", "bank", "bank_transfer", "bank transfer"):
        return "Net Banking"
    if mode_strip in ("card", "credit", "debit", "credit_card", "debit_card", "pos"):
        return "Card"
    if mode_strip in ("cheque", "check", "chq"):
        return "Cheque"
        
    for standard in ("Cash", "UPI", "Net Banking", "Card", "Cheque", "Other"):
        if mode_strip == standard.lower().replace(" ", "") or mode_strip == standard.lower():
            return standard
            
    return "Other"


def _normalize_name(name: Optional[str]) -> str:
    if not name:
        return ""
    return " ".join(str(name).split()).lower()


def _extract_payment_mode(record: dict) -> str:
    # 1. Try record's payment_mode
    mode = record.get("payment_mode")
    if mode:
        return _normalize_payment_mode(mode)
        
    # 2. Try nested payment payment_mode
    pay = record.get("payments") or record.get("payment") or {}
    if isinstance(pay, list) and pay:
        pay = pay[0]
    if isinstance(pay, dict):
        mode = pay.get("payment_mode")
        if mode:
            return _normalize_payment_mode(mode)
            
    # 3. Try parsing from reminder_note or notes
    note = record.get("reminder_note") or ""
    if not note and isinstance(pay, dict):
        note = pay.get("reminder_note") or ""
    if not note:
        note = record.get("notes") or ""
        
    import re
    match = re.search(r"\[Method:\s*([^\]\s]+)", note, re.IGNORECASE)
    if match:
        return _normalize_payment_mode(match.group(1))
        
    return "Other"


def _is_in_date_range(dt_str: Optional[str], date_from: Optional[str], date_to: Optional[str]) -> bool:
    if not dt_str or dt_str == "N/A":
        return False
    date_part = dt_str[:10]
    if date_from and date_part < date_from:
        return False
    if date_to and date_part > date_to:
        return False
    return True


def _get_aggregated_customers(customers: list[dict], payments: list[dict], transactions: list[dict], reminders: list[dict]) -> list[dict]:
    grouped = {}
    id_to_key = {}
    mobile_to_key = {}
    name_to_key = {}
    
    def find_existing_key(c_id, mobile, name):
        if c_id and str(c_id) in id_to_key:
            return id_to_key[str(c_id)]
        if mobile and str(mobile).strip() and str(mobile).strip() != "N/A":
            m_clean = str(mobile).strip()
            if m_clean in mobile_to_key:
                return mobile_to_key[m_clean]
        if name and str(name).strip() and str(name).strip().lower() != "n/a":
            n_clean = _normalize_name(name)
            if n_clean in name_to_key:
                return name_to_key[n_clean]
        return None

    def register_keys(key, c_id, mobile, name):
        if c_id:
            id_to_key[str(c_id)] = key
        if mobile and str(mobile).strip() and str(mobile).strip() != "N/A":
            mobile_to_key[str(mobile).strip()] = key
        if name and str(name).strip() and str(name).strip().lower() != "n/a":
            name_to_key[_normalize_name(name)] = key

    # Build a lookup of average total_amount per battery model name
    model_prices = {}
    for p in payments:
        b = p.get("batteries") or p.get("battery")
        if isinstance(b, list) and b:
            b = b[0]
        if isinstance(b, dict):
            model = (b.get("model_number") or b.get("model_name") or "").strip().upper()
            amt = float(p.get("total_amount") or 0.0)
            if model and amt > 0:
                model_prices.setdefault(model, []).append(amt)
                
    model_fallback_price = {}
    for model, prices in model_prices.items():
        model_fallback_price[model] = round(sum(prices) / len(prices), 2)

    # First pass: seed groups with customers
    for c in customers:
        c_id = c.get("id")
        mobile = c.get("mobile")
        name = c.get("name")
        
        key = find_existing_key(c_id, mobile, name)
        if not key:
            key = f"group_{c_id or len(grouped)}"
            grouped[key] = {
                "id": c_id,
                "name": name or "N/A",
                "mobile": mobile or "N/A",
                "vehicle_no": c.get("vehicle_no") or "",
                "vehicle_type": c.get("vehicle_type") or "",
                "area": c.get("area") or "",
                "pincode": c.get("pincode") or "",
                "purchase_type": c.get("purchase_type") or "",
                "created_at": c.get("created_at") or "",
                "scrap_expected_value": float(c.get("scrap_expected_value") or 0.0),
                "scrap_received_value": float(c.get("scrap_received_value") or 0.0),
                "scrap_payment_mode": c.get("scrap_payment_mode") or "",
                "scrap_received_date": c.get("scrap_received_date") or "",
                "batteries": [],
                "payments": [],
                "transactions": [],
                "reminders": []
            }
        
        register_keys(key, c_id, mobile, name)
        g = grouped[key]
        
        # Merge basic fields
        for f in ["vehicle_no", "vehicle_type", "area", "pincode", "purchase_type"]:
            if not g[f] and c.get(f):
                g[f] = c[f]
        if not g["created_at"] or (c.get("created_at") and c["created_at"] < g["created_at"]):
            g["created_at"] = c.get("created_at")
            
        if c.get("scrap_received_date"):
            if not g["scrap_received_date"] or c["scrap_received_date"] > g["scrap_received_date"]:
                g["scrap_received_date"] = c["scrap_received_date"]
                g["scrap_payment_mode"] = c.get("scrap_payment_mode") or ""
            g["scrap_received_value"] = max(g["scrap_received_value"], float(c.get("scrap_received_value") or 0.0))
            g["scrap_expected_value"] = max(g["scrap_expected_value"], float(c.get("scrap_expected_value") or 0.0))
            
        for b in (c.get("batteries") or []):
            if b not in g["batteries"]:
                g["batteries"].append(b)

    # Second pass: map payments, transactions, reminders
    for p in payments:
        cid = p.get("customer_id")
        cust_metadata = p.get("customers") or p.get("customer") or {}
        p_mobile = p.get("customer_mobile") or cust_metadata.get("mobile")
        p_name = p.get("customer_name") or cust_metadata.get("name")
        
        key = find_existing_key(cid, p_mobile, p_name)
        if not key:
            key = f"group_p_{len(grouped)}"
            grouped[key] = {
                "id": cid,
                "name": p_name or "N/A",
                "mobile": p_mobile or "N/A",
                "vehicle_no": "",
                "vehicle_type": "",
                "area": "",
                "pincode": "",
                "purchase_type": "",
                "created_at": p.get("created_at") or "",
                "scrap_expected_value": 0.0,
                "scrap_received_value": 0.0,
                "scrap_payment_mode": "",
                "scrap_received_date": "",
                "batteries": [],
                "payments": [],
                "transactions": [],
                "reminders": []
            }
        register_keys(key, cid, p_mobile, p_name)
        grouped[key]["payments"].append(p)
        
        # Add payment's nested battery if missing
        b = p.get("batteries") or p.get("battery")
        if isinstance(b, list) and b:
            b = b[0]
        if isinstance(b, dict):
            if b not in grouped[key]["batteries"]:
                grouped[key]["batteries"].append(b)
        
    for tx in transactions:
        cid = tx.get("customer_id")
        cust_metadata = tx.get("customers") or tx.get("customer") or {}
        tx_mobile = cust_metadata.get("mobile")
        tx_name = cust_metadata.get("name")
        
        key = find_existing_key(cid, tx_mobile, tx_name)
        if not key:
            key = f"group_tx_{len(grouped)}"
            grouped[key] = {
                "id": cid,
                "name": tx_name or "N/A",
                "mobile": tx_mobile or "N/A",
                "vehicle_no": "",
                "vehicle_type": "",
                "area": "",
                "pincode": "",
                "purchase_type": "",
                "created_at": tx.get("created_at") or "",
                "scrap_expected_value": 0.0,
                "scrap_received_value": 0.0,
                "scrap_payment_mode": "",
                "scrap_received_date": "",
                "batteries": [],
                "payments": [],
                "transactions": [],
                "reminders": []
            }
        register_keys(key, cid, tx_mobile, tx_name)
        grouped[key]["transactions"].append(tx)
        
        # Add transaction's payment's nested battery if missing
        pay = tx.get("payments") or tx.get("payment") or {}
        if isinstance(pay, list) and pay:
            pay = pay[0]
        if isinstance(pay, dict):
            b = pay.get("batteries") or pay.get("battery")
            if isinstance(b, list) and b:
                b = b[0]
            if isinstance(b, dict):
                if b not in grouped[key]["batteries"]:
                    grouped[key]["batteries"].append(b)
        
    for r in reminders:
        cid = r.get("customer_id")
        r_mobile = r.get("mobile_number")
        r_name = r.get("customer_name")
        
        key = find_existing_key(cid, r_mobile, r_name)
        if not key:
            key = f"group_r_{len(grouped)}"
            grouped[key] = {
                "id": cid,
                "name": r_name or "N/A",
                "mobile": r_mobile or "N/A",
                "vehicle_no": "",
                "vehicle_type": "",
                "area": "",
                "pincode": "",
                "purchase_type": "",
                "created_at": r.get("created_at") or "",
                "scrap_expected_value": 0.0,
                "scrap_received_value": 0.0,
                "scrap_payment_mode": "",
                "scrap_received_date": "",
                "batteries": [],
                "payments": [],
                "transactions": [],
                "reminders": []
            }
        register_keys(key, cid, r_mobile, r_name)
        grouped[key]["reminders"].append(r)

    # Third pass: build aggregates
    aggregated = []
    today_str = date.today().isoformat()
    
    for key, g in grouped.items():
        # Clean list duplicates
        g["payments"] = list({p["id"]: p for p in g["payments"] if p.get("id")}.values())
        g["transactions"] = list({t["id"]: t for t in g["transactions"] if t.get("id")}.values())
        g["reminders"] = list({r["id"]: r for r in g["reminders"] if r.get("id")}.values())
        
        # Battery models
        models = []
        for b in g["batteries"]:
            model = b.get("model_number") or b.get("model_name")
            if model:
                models.append(model)
        for p in g["payments"]:
            b = p.get("batteries") or p.get("battery")
            if isinstance(b, list) and b:
                b = b[0]
            if isinstance(b, dict):
                model = b.get("model_number") or b.get("model_name")
                if model:
                    models.append(model)
                    
        unique_models = sorted(list(set(models)))
        g["battery_models_str"] = ", ".join(unique_models) if unique_models else "N/A"
        
        # Latest Purchase Date
        sale_dates = []
        for b in g["batteries"]:
            if b.get("sale_date"):
                sale_dates.append(b["sale_date"][:10])
        for p in g["payments"]:
            b = p.get("batteries") or p.get("battery")
            if isinstance(b, list) and b:
                b = b[0]
            if isinstance(b, dict) and b.get("sale_date"):
                sale_dates.append(b["sale_date"][:10])
        g["latest_purchase_date"] = max(sale_dates) if sale_dates else "N/A"
        
        # Warranty Status
        w_expiry_dates = []
        for b in g["batteries"]:
            if b.get("warranty_expiry"):
                w_expiry_dates.append(b["warranty_expiry"][:10])
        for p in g["payments"]:
            b = p.get("batteries") or p.get("battery")
            if isinstance(b, list) and b:
                b = b[0]
            if isinstance(b, dict) and b.get("warranty_expiry"):
                w_expiry_dates.append(b["warranty_expiry"][:10])
        any_active = any(exp >= today_str for exp in w_expiry_dates if exp)
        g["warranty_status"] = "Active" if any_active else ("Expired" if w_expiry_dates else "N/A")
        
        # Sort transactions chronologically (oldest first) to compute running balance
        g_txs_sorted = sorted(g["transactions"], key=lambda x: x.get("created_at") or "")
        bal = 0.0
        for tx in g_txs_sorted:
            tx["recovered_payment_mode"] = _extract_payment_mode(tx)
            amt = float(tx.get("amount") or 0.0)
            if tx.get("transaction_type") == "ADDITION":
                bal += amt
            elif tx.get("transaction_type") == "PAYMENT":
                bal -= amt
            tx["running_balance"] = round(bal, 2)

        # Financial Totals & Validation
        pay_additions = sum(float(p.get("total_amount") or 0.0) for p in g["payments"])
        pay_payments = sum(float(p.get("paid_amount") or 0.0) for p in g["payments"])
        
        tx_additions = sum(float(tx.get("amount") or 0.0) for tx in g["transactions"] if tx.get("transaction_type") == "ADDITION")
        tx_payments = sum(float(tx.get("amount") or 0.0) for tx in g["transactions"] if tx.get("transaction_type") == "PAYMENT")
        
        total_bill = max(pay_additions, tx_additions)
        total_paid = max(pay_payments, tx_payments)
        
        # Case B fallback for customers who bought batteries and paid immediately
        if total_bill == 0.0 and g["batteries"]:
            inferred_bill = 0.0
            for b in g["batteries"]:
                model = (b.get("model_number") or b.get("model_name") or "").strip().upper()
                price = model_fallback_price.get(model)
                if not price:
                    b_type = (b.get("battery_type") or "").strip().upper()
                    if b_type == "2W":
                        price = 1500.0
                    elif b_type == "4W":
                        price = 5000.0
                    else:
                        price = 3000.0
                inferred_bill += price
            total_bill = inferred_bill
            total_paid = inferred_bill
            
        # Strict Business validation: Total Bill >= Total Paid
        if total_paid > total_bill:
            total_bill = total_paid
            
        outstanding = round(total_bill - total_paid, 2)
        if outstanding < 0:
            outstanding = 0.0
            
        g["total_additions"] = total_bill
        g["total_payments"] = total_paid
        g["outstanding_balance"] = outstanding
        
        # Payment Modes Used
        modes = []
        for p in g["payments"]:
            mode_p = p.get("payment_mode") or _extract_payment_mode(p)
            if mode_p:
                modes.append(mode_p)
        for tx in g["transactions"]:
            mode_t = tx.get("recovered_payment_mode") or _extract_payment_mode(tx)
            if mode_t:
                modes.append(mode_t)
        unique_modes = sorted(list(set(modes)))
        g["payment_modes_str"] = ", ".join(unique_modes) if unique_modes else "N/A"
        
        # Last Payment Mode & Date
        payment_txs = [tx for tx in g["transactions"] if tx.get("transaction_type") == "PAYMENT" and tx.get("created_at")]
        if payment_txs:
            payment_txs_sorted = sorted(payment_txs, key=lambda x: x.get("created_at") or "")
            last_pay_tx = payment_txs_sorted[-1]
            g["last_payment_mode"] = last_pay_tx.get("recovered_payment_mode") or _extract_payment_mode(last_pay_tx)
            g["last_payment_date"] = str(last_pay_tx.get("created_at"))[:10]
        else:
            # Fallback to payments table
            pay_modes = []
            for p in g["payments"]:
                if float(p.get("paid_amount") or 0.0) > 0:
                    pay_modes.append(p.get("payment_mode") or _extract_payment_mode(p))
            g["last_payment_mode"] = pay_modes[-1] if pay_modes else "N/A"
            
            pay_dates = [p.get("updated_at") or p.get("created_at") for p in g["payments"] if float(p.get("paid_amount") or 0.0) > 0]
            g["last_payment_date"] = str(max(pay_dates))[:10] if pay_dates else "N/A"
            
        # Last Activity Date
        activity_dates = []
        if g["created_at"]:
            activity_dates.append(g["created_at"][:10])
        for b in g["batteries"]:
            if b.get("sale_date"):
                activity_dates.append(b["sale_date"][:10])
        for p in g["payments"]:
            if p.get("created_at"):
                activity_dates.append(p["created_at"][:10])
            if p.get("updated_at"):
                activity_dates.append(p["updated_at"][:10])
        for tx in g["transactions"]:
            if tx.get("created_at"):
                activity_dates.append(tx["created_at"][:10])
        for r in g["reminders"]:
            if r.get("created_at"):
                activity_dates.append(r["created_at"][:10])
                
        g["last_activity_date"] = max(activity_dates) if activity_dates else "N/A"
        
        # Consolidated Address
        addr_parts = []
        if g["area"]:
            addr_parts.append(g["area"])
        if g["pincode"]:
            addr_parts.append(g["pincode"])
        g["consolidated_address"] = ", ".join(addr_parts) if addr_parts else "N/A"
        
        # Oldest unpaid date
        unpaid_dates = [p.get("created_at") for p in g["payments"] if not p.get("is_settled") and p.get("created_at")]
        g["oldest_unpaid_date"] = min(unpaid_dates)[:10] if unpaid_dates else None
        
        uncompleted_rems = [r.get("reminder_date") for r in g["reminders"] if not r.get("is_completed") and r.get("reminder_date")]
        g["due_date"] = min(uncompleted_rems)[:10] if uncompleted_rems else "N/A"
        
        aggregated.append(g)
        
    return aggregated


def _write_summary_cards(ws, title: str, subtitle: str, cards: list[dict]) -> None:
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color="1E3A5F")
    ws.cell(row=2, column=1, value=subtitle).font = Font(italic=True, size=9, color="555555")
    
    for idx, card in enumerate(cards):
        col = idx * 2 + 2  # Columns B, D, F, H, J, etc.
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        
        lbl_cell = ws.cell(row=4, column=col, value=card["label"])
        lbl_cell.font = Font(bold=True, size=9, color="555555")
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
        lbl_cell.fill = CARD_LABEL_FILL
        
        val_cell = ws.cell(row=5, column=col, value=card["value"])
        val_cell.font = Font(bold=True, size=12, color="1E3A5F")
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.fill = CARD_VALUE_FILL
        
        if card.get("is_currency"):
            val_cell.number_format = '"₹"#,##0.00'
        elif isinstance(card["value"], (int, float)):
            val_cell.number_format = '#,##0'


def _draw_dashboard_card(ws, start_col: int, start_row: int, label: str, value, is_currency: bool = False) -> None:
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)
    ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 2, end_column=start_col + 1)
    
    lbl_cell = ws.cell(row=start_row, column=start_col, value=label)
    lbl_cell.font = Font(bold=True, size=10, color="FFFFFF")
    lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
    lbl_cell.fill = HEADER_FILL
    
    val_cell = ws.cell(row=start_row + 1, column=start_col, value=value)
    val_cell.font = Font(bold=True, size=20, color="1E3A5F")
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    val_cell.fill = CARD_VALUE_FILL
    
    if is_currency:
        val_cell.number_format = '"₹"#,##0.00'
    else:
        val_cell.number_format = '#,##0'


def _apply_table_formatting(ws, start_row: int, headers: list[str], currency_cols: list[int] = [], date_cols: list[int] = [], number_cols: list[int] = []) -> None:
    # 1. Headers
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        
    # 2. Zebra rows and formats
    last_row = ws.max_row
    for r in range(start_row + 1, last_row + 1):
        is_even = (r - start_row) % 2 == 0
        row_fill = ZEBRA_FILL if is_even else None
        
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            if row_fill and cell.fill.fill_type is None:  # don't overwrite custom status fills
                cell.fill = row_fill
                
            if c in currency_cols:
                cell.number_format = '"₹"#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c in date_cols:
                cell.number_format = 'yyyy-mm-dd'
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c in number_cols:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                if cell.alignment.horizontal is None:
                    # Default left alignment for text columns
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
    # 3. Columns width auto-fit (measuring table rows only)
    for col in ws.columns:
        vals = []
        for cell in col:
            if cell.row >= start_row:
                vals.append(str(cell.value or ""))
        max_len = max((len(v) for v in vals), default=10)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)
        
    # 4. Freeze Header Row
    ws.freeze_panes = f"A{start_row + 1}"
    
    # 5. Dropdown auto filters
    if last_row > start_row:
        col_letter = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A{start_row}:{col_letter}{last_row}"


# ---------------------------------------------------------------------------
# Redesigned Sheet Builders
# ---------------------------------------------------------------------------

def _build_business_summary_sheet(ws, customers_rows, batteries_rows, payments_rows, scrap_rows, shops_rows, stock_rows, transactions_rows=None) -> None:
    ws.sheet_view.showGridLines = True
    
    ws.cell(row=2, column=2, value="SHREE GANADHISH BATTERY SERVICES").font = Font(bold=True, size=16, color="1E3A5F")
    ws.cell(row=3, column=2, value="MANAGEMENT BUSINESS SUMMARY REPORT").font = Font(bold=True, size=11, color="555555")
    ws.cell(row=4, column=2, value=f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = Font(italic=True, size=9, color="777777")
    
    total_customers = len(customers_rows)
    total_batteries = len(batteries_rows)
    total_outstanding = sum(float(p.get("pending_amount") or 0.0) for p in payments_rows)
    total_scrap = len(scrap_rows)
    total_shops = len(shops_rows)
    total_stock = sum(int(st.get("quantity") or 0) for st in stock_rows)
    
    _draw_dashboard_card(ws, start_col=2, start_row=6, label="TOTAL CUSTOMERS", value=total_customers)
    _draw_dashboard_card(ws, start_col=5, start_row=6, label="BATTERIES SOLD", value=total_batteries)
    _draw_dashboard_card(ws, start_col=2, start_row=9, label="OUTSTANDING UDHARI", value=total_outstanding, is_currency=True)
    _draw_dashboard_card(ws, start_col=5, start_row=9, label="SCRAP BATTERIES COLLECTED", value=total_scrap)
    _draw_dashboard_card(ws, start_col=2, start_row=12, label="WHOLESALE SHOPS", value=total_shops)
    _draw_dashboard_card(ws, start_col=5, start_row=12, label="TOTAL STOCK INVENTORY", value=total_stock)
    
    # Calculate collections from transactions
    collections = {
        "Cash": 0.0,
        "UPI": 0.0,
        "Net Banking": 0.0,
        "Card": 0.0,
        "Cheque": 0.0,
        "Other": 0.0
    }
    if transactions_rows:
        for tx in transactions_rows:
            if tx.get("transaction_type") == "PAYMENT":
                mode = tx.get("recovered_payment_mode") or _extract_payment_mode(tx)
                if mode in collections:
                    collections[mode] += float(tx.get("amount") or 0.0)
                else:
                    collections["Other"] += float(tx.get("amount") or 0.0)
                    
    _draw_dashboard_card(ws, start_col=8, start_row=6, label="CASH COLLECTION", value=collections["Cash"], is_currency=True)
    _draw_dashboard_card(ws, start_col=11, start_row=6, label="CARD COLLECTION", value=collections["Card"], is_currency=True)
    _draw_dashboard_card(ws, start_col=8, start_row=9, label="UPI COLLECTION", value=collections["UPI"], is_currency=True)
    _draw_dashboard_card(ws, start_col=11, start_row=9, label="CHEQUE COLLECTION", value=collections["Cheque"], is_currency=True)
    _draw_dashboard_card(ws, start_col=8, start_row=12, label="NET BANKING", value=collections["Net Banking"], is_currency=True)
    _draw_dashboard_card(ws, start_col=11, start_row=12, label="OTHER COLLECTION", value=collections["Other"], is_currency=True)
    
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 5
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 5
    ws.column_dimensions["K"].width = 20
    ws.column_dimensions["L"].width = 20



def _build_customers_sheet(ws, aggregated_customers: list[dict]) -> None:
    total_customers = len(aggregated_customers)
    total_outstanding = sum(c["outstanding_balance"] for c in aggregated_customers)
    
    cards = [
        {"label": "Total Customers", "value": total_customers},
        {"label": "Total Outstanding Balance", "value": total_outstanding, "is_currency": True}
    ]
    _write_summary_cards(ws, "CUSTOMER DIRECTORY EXPORT", "Alphabetical listing of all customers with contact details and warranty summary", cards)
    
    headers = [
        "Customer Name", "Mobile", "Village / Address", "Battery Model",
        "Purchase Date", "Warranty Status", "Outstanding Balance", "Last Activity Date"
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=7, column=col_idx, value=h)
        
    sorted_customers = sorted(aggregated_customers, key=lambda x: (x["name"] or "").strip().lower())
    for r_idx, c in enumerate(sorted_customers, 8):
        ws.cell(row=r_idx, column=1, value=c["name"])
        ws.cell(row=r_idx, column=2, value=c["mobile"])
        ws.cell(row=r_idx, column=3, value=c["consolidated_address"])
        ws.cell(row=r_idx, column=4, value=c["battery_models_str"])
        ws.cell(row=r_idx, column=5, value=c["latest_purchase_date"])
        ws.cell(row=r_idx, column=6, value=c["warranty_status"])
        ws.cell(row=r_idx, column=7, value=c["outstanding_balance"])
        ws.cell(row=r_idx, column=8, value=c["last_activity_date"])
        
    _apply_table_formatting(ws, start_row=7, headers=headers, currency_cols=[7], date_cols=[5, 8])


def _build_batteries_sheet(ws, batteries_rows: list[dict]) -> None:
    today = date.today()
    processed_batteries = []
    active_count = 0
    expired_count = 0
    soon_count = 0
    
    for r in batteries_rows:
        customer = r.get("customers") or r.get("customer") or {}
        name = customer.get("name", "N/A")
        mobile = customer.get("mobile", "N/A")
        model = r.get("model_number") or r.get("model_name") or "N/A"
        serial = r.get("serial_number") or "N/A"
        sale_d = r.get("sale_date")
        sale_date_str = sale_d[:10] if sale_d else "N/A"
        w_months = r.get("warranty_months") or 0
        w_period = f"{w_months} Months" if w_months else "N/A"
        
        w_exp = r.get("warranty_expiry")
        w_exp_str = w_exp[:10] if w_exp else "N/A"
        
        status = "Expired"
        remaining = "Expired"
        sort_key = 3  # 1: Expiring Soon, 2: Active, 3: Expired
        
        if w_exp:
            try:
                exp_date = date.fromisoformat(w_exp[:10])
                delta = (exp_date - today).days
                if delta < 0:
                    status = "Expired"
                    remaining = "Expired"
                    sort_key = 3
                    expired_count += 1
                elif delta <= 30:
                    status = "Expiring Soon"
                    remaining = f"{delta} Days"
                    sort_key = 1
                    soon_count += 1
                else:
                    status = "Active"
                    remaining = f"{delta} Days"
                    sort_key = 2
                    active_count += 1
            except Exception:
                expired_count += 1
        else:
            expired_count += 1
            
        processed_batteries.append({
            "name": name,
            "mobile": mobile,
            "model": model,
            "serial": serial,
            "sale_date": sale_date_str,
            "period": w_period,
            "expiry_date": w_exp_str,
            "remaining": remaining,
            "status": status,
            "sort_key": sort_key
        })
        
    sorted_batteries = sorted(processed_batteries, key=lambda x: (x["sort_key"], x["expiry_date"]))
    
    cards = [
        {"label": "Active Warranties", "value": active_count},
        {"label": "Expiring Soon", "value": soon_count},
        {"label": "Expired Warranties", "value": expired_count}
    ]
    _write_summary_cards(ws, "WARRANTY REGISTRY EXPORT", "Tracking warranty statuses of sold customer batteries", cards)
    
    headers = [
        "Customer Name", "Mobile", "Battery Model", "Serial Number",
        "Purchase Date", "Warranty Period", "Warranty Expiry Date", "Warranty Remaining", "Status"
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=7, column=col_idx, value=h)
        
    for r_idx, b in enumerate(sorted_batteries, 8):
        ws.cell(row=r_idx, column=1, value=b["name"])
        ws.cell(row=r_idx, column=2, value=b["mobile"])
        ws.cell(row=r_idx, column=3, value=b["model"])
        ws.cell(row=r_idx, column=4, value=b["serial"])
        ws.cell(row=r_idx, column=5, value=b["sale_date"])
        ws.cell(row=r_idx, column=6, value=b["period"])
        ws.cell(row=r_idx, column=7, value=b["expiry_date"])
        ws.cell(row=r_idx, column=8, value=b["remaining"])
        
        status_cell = ws.cell(row=r_idx, column=9, value=b["status"])
        if b["status"] == "Expiring Soon":
            status_cell.fill = YELLOW_FILL
        elif b["status"] == "Expired":
            status_cell.fill = RED_FILL
        elif b["status"] == "Active":
            status_cell.fill = GREEN_FILL
            
    _apply_table_formatting(ws, start_row=7, headers=headers, date_cols=[5, 7])


def _build_customer_summary_sheet(ws, aggregated_customers: list[dict]) -> None:
    # Case A: Exclude customers with no business activity
    active_customers = [
        c for c in aggregated_customers
        if len(c.get("batteries") or []) > 0 or len(c.get("payments") or []) > 0 or len(c.get("transactions") or []) > 0
    ]
    
    total_additions = sum(c["total_additions"] for c in active_customers)
    total_payments = sum(c["total_payments"] for c in active_customers)
    total_outstanding = sum(c["outstanding_balance"] for c in active_customers)
    
    collections = {
        "Cash": 0.0,
        "UPI": 0.0,
        "Net Banking": 0.0,
        "Card": 0.0,
        "Cheque": 0.0,
        "Other": 0.0
    }
    
    for c in active_customers:
        for tx in c["transactions"]:
            if tx.get("transaction_type") == "PAYMENT":
                mode = tx.get("recovered_payment_mode") or _extract_payment_mode(tx)
                if mode in collections:
                    collections[mode] += float(tx.get("amount") or 0.0)
                else:
                    collections["Other"] += float(tx.get("amount") or 0.0)
                    
    cards = [
        {"label": "Total Bill", "value": total_additions, "is_currency": True},
        {"label": "Total Paid", "value": total_payments, "is_currency": True},
        {"label": "Outstanding", "value": total_outstanding, "is_currency": True},
        {"label": "Cash Collection", "value": collections["Cash"], "is_currency": True},
        {"label": "UPI Collection", "value": collections["UPI"], "is_currency": True},
        {"label": "Net Banking Collection", "value": collections["Net Banking"], "is_currency": True},
        {"label": "Card Collection", "value": collections["Card"], "is_currency": True},
        {"label": "Cheque Collection", "value": collections["Cheque"], "is_currency": True}
    ]
    _write_summary_cards(ws, "CUSTOMER SUMMARY", "Financial statement and payment summaries per customer ledger", cards)
    
    headers = [
        "Customer Name", "Battery Models", "Total Bill Amount", "Total Paid Amount",
        "Outstanding Balance", "Payment Modes Used", "Last Payment Mode", "Last Payment Date"
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=7, column=col_idx, value=h)
        
    sorted_customers = sorted(active_customers, key=lambda x: (x["name"] or "").strip().lower())
    for r_idx, c in enumerate(sorted_customers, 8):
        ws.cell(row=r_idx, column=1, value=c["name"])
        ws.cell(row=r_idx, column=2, value=c["battery_models_str"])
        ws.cell(row=r_idx, column=3, value=c["total_additions"])
        ws.cell(row=r_idx, column=4, value=c["total_payments"])
        ws.cell(row=r_idx, column=5, value=c["outstanding_balance"])
        ws.cell(row=r_idx, column=6, value=c["payment_modes_str"])
        ws.cell(row=r_idx, column=7, value=c.get("last_payment_mode") or "N/A")
        ws.cell(row=r_idx, column=8, value=c.get("last_payment_date") or "N/A")
        
    _apply_table_formatting(ws, start_row=7, headers=headers, currency_cols=[3, 4, 5], date_cols=[8])


def _build_customer_transactions_sheet(ws, tx_rows: list[dict]) -> None:
    total_tx = len(tx_rows)
    total_payments = sum(float(tx.get("amount") or 0.0) for tx in tx_rows if tx.get("transaction_type") == "PAYMENT")
    total_additions = sum(float(tx.get("amount") or 0.0) for tx in tx_rows if tx.get("transaction_type") == "ADDITION")
    
    cards = [
        {"label": "Total Transactions", "value": total_tx},
        {"label": "Total Additions Amount", "value": total_additions, "is_currency": True},
        {"label": "Total Payments Amount", "value": total_payments, "is_currency": True}
    ]
    _write_summary_cards(ws, "TRANSACTION HISTORY", "Detailed history of wholesale and retail payment transactions", cards)
    
    headers = [
        "Date", "Customer Name", "Battery Model", "Transaction Type",
        "Amount", "Payment Mode", "Notes", "Running Balance"
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=7, column=col_idx, value=h)
        
    sorted_tx = sorted(tx_rows, key=lambda x: x.get("created_at") or "", reverse=True)
    for r_idx, tx in enumerate(sorted_tx, 8):
        customer = tx.get("customers") or tx.get("customer") or {}
        payment = tx.get("payments") or tx.get("payment") or {}
        # Support if payment is list
        if isinstance(payment, list) and payment:
            payment = payment[0]
        elif not isinstance(payment, dict):
            payment = {}
            
        battery = payment.get("batteries") or payment.get("battery") or {}
        if isinstance(battery, list) and battery:
            battery = battery[0]
        elif not isinstance(battery, dict):
            battery = {}
            
        battery_model = f"{battery.get('battery_type', '')} {battery.get('model_number', '')}".strip() or "N/A"
        if not battery_model or battery_model == "N/A":
            battery_model = battery.get("model_name") or "N/A"
            
        ws.cell(row=r_idx, column=1, value=str(tx.get("created_at", ""))[:10])
        ws.cell(row=r_idx, column=2, value=customer.get("name", "N/A"))
        ws.cell(row=r_idx, column=3, value=battery_model)
        ws.cell(row=r_idx, column=4, value=tx.get("transaction_type", "N/A"))
        ws.cell(row=r_idx, column=5, value=float(tx.get("amount") or 0.0))
        ws.cell(row=r_idx, column=6, value=tx.get("recovered_payment_mode") or _extract_payment_mode(tx))
        ws.cell(row=r_idx, column=7, value=tx.get("notes") or "")
        ws.cell(row=r_idx, column=8, value=tx.get("running_balance") or 0.0)
        
    _apply_table_formatting(ws, start_row=7, headers=headers, currency_cols=[5, 8], date_cols=[1])


def _build_payments_sheet(ws, aggregated_customers: list[dict]) -> None:
    today = date.today()
    today_str = today.isoformat()
    processed = []
    total_outstanding = 0.0
    overdue_count = 0
    pending_count = 0
    
    # Filter to active debtors (outstanding_balance > 0.01)
    active_debtors = [c for c in aggregated_customers if c["outstanding_balance"] > 0.01]
    
    collections = {
        "Cash": 0.0,
        "UPI": 0.0,
        "Net Banking": 0.0,
        "Card": 0.0,
        "Cheque": 0.0,
        "Other": 0.0
    }
    
    for c in active_debtors:
        outstanding = c["outstanding_balance"]
        total_outstanding += outstanding
        
        days_pending = 0
        oldest_unpaid = c.get("oldest_unpaid_date")
        if outstanding > 0 and oldest_unpaid:
            try:
                days_pending = (today - date.fromisoformat(oldest_unpaid)).days
            except Exception:
                pass
                
        # Calculate payment mode collections for active debtors
        for tx in c["transactions"]:
            if tx.get("transaction_type") == "PAYMENT":
                mode = tx.get("recovered_payment_mode") or _extract_payment_mode(tx)
                if mode in collections:
                    collections[mode] += float(tx.get("amount") or 0.0)
                else:
                    collections["Other"] += float(tx.get("amount") or 0.0)
                    
        due_date = c["due_date"]
        is_overdue = False
        if outstanding > 0:
            if days_pending > 30:
                is_overdue = True
            elif due_date != "N/A" and due_date < today_str:
                is_overdue = True
                
        if is_overdue:
            status = "Overdue"
            overdue_count += 1
        else:
            status = "Pending"
            pending_count += 1
            
        processed.append({
            "name": c["name"],
            "mobile": c["mobile"],
            "total_bill": c["total_additions"],
            "total_paid": c["total_payments"],
            "outstanding": outstanding,
            "modes": c["payment_modes_str"],
            "last_payment_mode": c.get("last_payment_mode") or "N/A",
            "last_payment_date": c.get("last_payment_date") or "N/A",
            "due_date": due_date,
            "days_pending": days_pending,
            "status": status
        })
        
    sorted_processed = sorted(processed, key=lambda x: x["outstanding"], reverse=True)
    
    cards = [
        {"label": "Total Outstanding", "value": total_outstanding, "is_currency": True},
        {"label": "Pending Accounts", "value": pending_count},
        {"label": "Overdue Accounts", "value": overdue_count},
        {"label": "Cash Collection", "value": collections["Cash"], "is_currency": True},
        {"label": "UPI Collection", "value": collections["UPI"], "is_currency": True},
        {"label": "Net Banking Collection", "value": collections["Net Banking"], "is_currency": True},
        {"label": "Card Collection", "value": collections["Card"], "is_currency": True},
        {"label": "Cheque Collection", "value": collections["Cheque"], "is_currency": True}
    ]
    _write_summary_cards(ws, "UDHARI LEDGER", "Debt recovery and outstanding collection tracking report", cards)
    
    headers = [
        "Customer Name", "Mobile", "Total Bill", "Total Paid", "Outstanding",
        "Payment Modes Used", "Last Payment Mode", "Last Payment Date", "Due Date", "Days Pending"
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=7, column=col_idx, value=h)
        
    for r_idx, p in enumerate(sorted_processed, 8):
        ws.cell(row=r_idx, column=1, value=p["name"])
        ws.cell(row=r_idx, column=2, value=p["mobile"])
        ws.cell(row=r_idx, column=3, value=p["total_bill"])
        ws.cell(row=r_idx, column=4, value=p["total_paid"])
        
        out_cell = ws.cell(row=r_idx, column=5, value=p["outstanding"])
        if p["status"] == "Overdue":
            out_cell.fill = RED_FILL
        else:
            out_cell.fill = YELLOW_FILL
            
        ws.cell(row=r_idx, column=6, value=p["modes"])
        ws.cell(row=r_idx, column=7, value=p["last_payment_mode"])
        ws.cell(row=r_idx, column=8, value=p["last_payment_date"])
        ws.cell(row=r_idx, column=9, value=p["due_date"])
        ws.cell(row=r_idx, column=10, value=p["days_pending"])
        
    _apply_table_formatting(ws, start_row=7, headers=headers, currency_cols=[3, 4, 5], date_cols=[8, 9], number_cols=[10])


def _build_scrap_payments_sheet(ws, aggregated_customers: list[dict]) -> None:
    scrap_customers = [c for c in aggregated_customers if c.get("scrap_received_date") and c["scrap_received_date"] != "N/A"]
    
    total_qty = 0
    total_value = 0.0
    cash_payments = 0.0
    upi_payments = 0.0
    other_payments = 0.0
    processed = []
    
    for c in scrap_customers:
        qty = 1
        val = c["scrap_received_value"]
        mode = _normalize_payment_mode(c["scrap_payment_mode"])
        
        total_qty += qty
        total_value += val
        if mode == "Cash":
            cash_payments += val
        elif mode == "UPI":
            upi_payments += val
        else:
            other_payments += val
            
        processed.append({
            "name": c["name"],
            "mobile": c["mobile"],
            "model": c["battery_models_str"],
            "qty": qty,
            "value": val,
            "mode": mode,
            "date": c["scrap_received_date"][:10],
            "collected_by": "Admin"
        })
        
    sorted_processed = sorted(processed, key=lambda x: x["date"], reverse=True)
    
    cards = [
        {"label": "Total Scrap Batteries", "value": total_qty},
        {"label": "Total Scrap Amount", "value": total_value, "is_currency": True},
        {"label": "Cash Payments", "value": cash_payments, "is_currency": True},
        {"label": "UPI Payments", "value": upi_payments, "is_currency": True},
        {"label": "Other Payments", "value": other_payments, "is_currency": True}
    ]
    _write_summary_cards(ws, "SCRAP BATTERIES", "Scrap battery collections log and payout summary breakdown", cards)
    
    headers = [
        "Customer Name", "Mobile", "Battery Model", "Scrap Battery Quantity",
        "Scrap Value", "Payment Mode", "Payment Date", "Collected By"
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=7, column=col_idx, value=h)
        
    for r_idx, s in enumerate(sorted_processed, 8):
        ws.cell(row=r_idx, column=1, value=s["name"])
        ws.cell(row=r_idx, column=2, value=s["mobile"])
        ws.cell(row=r_idx, column=3, value=s["model"])
        ws.cell(row=r_idx, column=4, value=s["qty"])
        ws.cell(row=r_idx, column=5, value=s["value"])
        ws.cell(row=r_idx, column=6, value=s["mode"])
        ws.cell(row=r_idx, column=7, value=s["date"])
        ws.cell(row=r_idx, column=8, value=s["collected_by"])
        
    _apply_table_formatting(ws, start_row=7, headers=headers, currency_cols=[5], date_cols=[7], number_cols=[4])


def _build_stock_sheet(ws, stock_rows: list[dict]) -> None:
    processed = []
    total_qty = 0
    out_count = 0
    low_count = 0
    
    for r in stock_rows:
        model_name = r.get("model_name", "").strip()
        parts = model_name.split(" ", 1)
        brand = parts[0] if parts else "N/A"
        model = parts[1] if len(parts) > 1 else brand
        
        qty = r.get("quantity", 0)
        threshold = r.get("low_stock_threshold", 2)
        total_qty += qty
        
        if qty == 0:
            status = "Out Of Stock"
            sort_key = 1
            out_count += 1
        elif qty <= threshold:
            status = "Low Stock"
            sort_key = 2
            low_count += 1
        else:
            status = "In Stock"
            sort_key = 3
            
        processed.append({
            "model": model,
            "brand": brand,
            "qty": qty,
            "threshold": threshold,
            "status": status,
            "sort_key": sort_key
        })
        
    sorted_processed = sorted(processed, key=lambda x: (x["sort_key"], x["brand"], x["model"]))
    
    cards = [
        {"label": "Total Stock Units", "value": total_qty},
        {"label": "Out of Stock Models", "value": out_count},
        {"label": "Low Stock Models", "value": low_count}
    ]
    _write_summary_cards(ws, "STOCK & INVENTORY", "Current warehouse inventory stock levels and threshold alerts", cards)
    
    headers = ["Battery Model", "Brand", "Quantity", "Low Stock Threshold", "Status"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=7, column=col_idx, value=h)
        
    for r_idx, st in enumerate(sorted_processed, 8):
        ws.cell(row=r_idx, column=1, value=st["model"])
        ws.cell(row=r_idx, column=2, value=st["brand"])
        ws.cell(row=r_idx, column=3, value=st["qty"])
        ws.cell(row=r_idx, column=4, value=st["threshold"])
        
        status_cell = ws.cell(row=r_idx, column=5, value=st["status"])
        if st["status"] == "Out Of Stock":
            status_cell.fill = RED_FILL
        elif st["status"] == "Low Stock":
            status_cell.fill = YELLOW_FILL
        elif st["status"] == "In Stock":
            status_cell.fill = GREEN_FILL
            
    _apply_table_formatting(ws, start_row=7, headers=headers, number_cols=[3, 4])


def _build_shops_sheet(ws, shops_rows: list[dict], purchases_rows: list[dict], payments_rows: list[dict], transactions_rows: list[dict]) -> None:
    purchases_by_shop = {}
    for p in purchases_rows:
        sid = p.get("shop_id")
        if sid:
            purchases_by_shop.setdefault(str(sid), []).append(p)
            
    payments_by_shop = {}
    for pm in payments_rows:
        sid = pm.get("shop_id")
        if sid:
            payments_by_shop.setdefault(str(sid), []).append(pm)
            
    tx_by_shop = {}
    for tx in transactions_rows:
        sid = tx.get("shop_id")
        if sid:
            tx_by_shop.setdefault(str(sid), []).append(tx)
            
    processed = []
    total_purchases_sum = 0.0
    total_payments_sum = 0.0
    total_outstanding_sum = 0.0
    
    for s in shops_rows:
        sid_str = str(s["id"])
        s_purchases = purchases_by_shop.get(sid_str, [])
        s_payments = payments_by_shop.get(sid_str, [])
        s_tx = tx_by_shop.get(sid_str, [])
        
        purchases_val = sum(float(p.get("amount") or 0.0) for p in s_purchases)
        outstanding_val = sum(float(pm.get("pending_amount") or 0.0) for pm in s_payments)
        payments_val = max(0.0, purchases_val - outstanding_val)
        
        total_purchases_sum += purchases_val
        total_payments_sum += payments_val
        total_outstanding_sum += outstanding_val
        
        dates = []
        if s.get("created_at"):
            dates.append(s["created_at"][:10])
        for p in s_purchases:
            if p.get("purchase_date"):
                dates.append(p["purchase_date"][:10])
        for pm in s_payments:
            if pm.get("created_at"):
                dates.append(pm["created_at"][:10])
        for tx in s_tx:
            if tx.get("created_at"):
                dates.append(tx["created_at"][:10])
        last_date = max(dates) if dates else "N/A"
        
        modes = []
        for p in s_purchases:
            if p.get("payment_mode"):
                modes.append(_normalize_payment_mode(p["payment_mode"]))
        for tx in s_tx:
            if tx.get("payment_mode"):
                modes.append(_normalize_payment_mode(tx["payment_mode"]))
        unique_modes = sorted(list(set(modes)))
        modes_str = ", ".join(unique_modes) if unique_modes else "N/A"
        
        processed.append({
            "name": s.get("shop_name", "N/A"),
            "owner": s.get("owner_name", "N/A"),
            "mobile": s.get("mobile", "N/A"),
            "purchases": purchases_val,
            "payments": payments_val,
            "outstanding": outstanding_val,
            "last_date": last_date,
            "modes": modes_str
        })
        
    sorted_processed = sorted(processed, key=lambda x: x["outstanding"], reverse=True)
    
    cards = [
        {"label": "Total Shops", "value": len(shops_rows)},
        {"label": "Total Purchases", "value": total_purchases_sum, "is_currency": True},
        {"label": "Total Payments", "value": total_payments_sum, "is_currency": True},
        {"label": "Total Outstanding", "value": total_outstanding_sum, "is_currency": True}
    ]
    _write_summary_cards(ws, "SHOPS / RETAILERS", "Directory of retailers, purchases, payments, and wholesaler balance", cards)
    
    headers = [
        "Shop Name", "Owner Name", "Mobile", "Total Purchases",
        "Total Payments", "Outstanding Balance", "Last Transaction Date", "Payment Modes Used"
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=7, column=col_idx, value=h)
        
    for r_idx, sp in enumerate(sorted_processed, 8):
        ws.cell(row=r_idx, column=1, value=sp["name"])
        ws.cell(row=r_idx, column=2, value=sp["owner"])
        ws.cell(row=r_idx, column=3, value=sp["mobile"])
        ws.cell(row=r_idx, column=4, value=sp["purchases"])
        ws.cell(row=r_idx, column=5, value=sp["payments"])
        ws.cell(row=r_idx, column=6, value=sp["outstanding"])
        ws.cell(row=r_idx, column=7, value=sp["last_date"])
        ws.cell(row=r_idx, column=8, value=sp["modes"])
        
    _apply_table_formatting(ws, start_row=7, headers=headers, currency_cols=[4, 5, 6], date_cols=[7])


def _build_shop_payment_transactions_sheet(ws, tx_rows: list[dict]) -> None:
    total_tx = len(tx_rows)
    total_payouts = sum(float(r.get("amount") or 0.0) for r in tx_rows)
    
    cards = [
        {"label": "Total Transactions", "value": total_tx},
        {"label": "Total Amount Transacted", "value": total_payouts, "is_currency": True}
    ]
    _write_summary_cards(ws, "SHOP TRANSACTIONS", "Chronological ledger log of wholesale payment transactions", cards)
    
    headers = ["Shop Name", "Transaction Type", "Amount", "Notes", "Payment Mode", "Payment Date"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=7, column=col_idx, value=h)
        
    sorted_tx = sorted(tx_rows, key=lambda x: x.get("created_at") or "", reverse=True)
    for r_idx, tx in enumerate(sorted_tx, 8):
        shop = tx.get("shops") or tx.get("shop") or {}
        ws.cell(row=r_idx, column=1, value=shop.get("shop_name", "N/A"))
        ws.cell(row=r_idx, column=2, value=tx.get("transaction_type", "N/A"))
        ws.cell(row=r_idx, column=3, value=float(tx.get("amount") or 0.0))
        ws.cell(row=r_idx, column=4, value=tx.get("notes") or "")
        ws.cell(row=r_idx, column=5, value=_normalize_payment_mode(tx.get("payment_mode")))
        ws.cell(row=r_idx, column=6, value=str(tx.get("created_at", ""))[:10])
        
    _apply_table_formatting(ws, start_row=7, headers=headers, currency_cols=[3], date_cols=[6])


def _build_reminders_sheet(ws, rows: list[dict]) -> None:
    total_rem = len(rows)
    comp_count = sum(1 for r in rows if r.get("is_completed"))
    pending_count = total_rem - comp_count
    
    cards = [
        {"label": "Total Reminders", "value": total_rem},
        {"label": "Completed", "value": comp_count},
        {"label": "Pending", "value": pending_count}
    ]
    _write_summary_cards(ws, "SERVICE REMINDERS EXPORT", "All service and warranty reminders with tracking history", cards)
    
    headers = [
        "Customer Name", "Mobile Number", "Battery Model", "Serial Number",
        "Battery Type", "Reminder Type", "Reminder Date", "Status",
        "Message Sent", "Sent At", "Completed", "Notes", "Created At"
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=7, column=col_idx, value=h)
        
    sorted_rows = sorted(rows, key=lambda x: x.get("reminder_date") or "")
    for r_idx, r in enumerate(sorted_rows, 8):
        ws.cell(row=r_idx, column=1, value=r.get("customer_name", ""))
        ws.cell(row=r_idx, column=2, value=r.get("mobile_number", ""))
        ws.cell(row=r_idx, column=3, value=r.get("battery_model", ""))
        ws.cell(row=r_idx, column=4, value=r.get("battery_serial", ""))
        ws.cell(row=r_idx, column=5, value=r.get("battery_type", ""))
        ws.cell(row=r_idx, column=6, value=r.get("reminder_type", ""))
        ws.cell(row=r_idx, column=7, value=str(r.get("reminder_date", ""))[:10])
        ws.cell(row=r_idx, column=8, value=r.get("reminder_status", ""))
        ws.cell(row=r_idx, column=9, value="Yes" if r.get("message_sent") else "No")
        ws.cell(row=r_idx, column=10, value=str(r.get("sent_at", ""))[:10] if r.get("sent_at") else "")
        ws.cell(row=r_idx, column=11, value="Yes" if r.get("is_completed") else "No")
        ws.cell(row=r_idx, column=12, value=r.get("notes", ""))
        ws.cell(row=r_idx, column=13, value=str(r.get("created_at", ""))[:10])
        
    _apply_table_formatting(ws, start_row=7, headers=headers, date_cols=[7, 10, 13])


def _build_shop_purchases_sheet(ws, rows: list[dict]) -> None:
    total_pur_val = sum(float(r.get("amount") or 0.0) for r in rows)
    total_udhari_val = sum(float(r.get("udhari_amount") or 0.0) for r in rows)
    
    cards = [
        {"label": "Total Purchases Value", "value": total_pur_val, "is_currency": True},
        {"label": "Outstanding Udhari Value", "value": total_udhari_val, "is_currency": True}
    ]
    _write_summary_cards(ws, "SHOP PURCHASES", "Chronological records of batteries purchased from wholesale shops", cards)
    
    headers = [
        "Shop Name", "Owner Name", "Mobile", "Battery Model", "Quantity", "Serial Numbers",
        "Invoice Number", "Stock Reduced", "Purchase Date", "Amount", "Udhari Amount",
        "Payment Mode", "Created At"
    ]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=7, column=col_idx, value=h)
        
    sorted_rows = sorted(rows, key=lambda x: x.get("purchase_date") or "", reverse=True)
    for r_idx, r in enumerate(sorted_rows, 8):
        shop = r.get("shops") or r.get("shop") or {}
        ws.cell(row=r_idx, column=1, value=shop.get("shop_name", ""))
        ws.cell(row=r_idx, column=2, value=shop.get("owner_name", ""))
        ws.cell(row=r_idx, column=3, value=shop.get("mobile", ""))
        ws.cell(row=r_idx, column=4, value=r.get("battery_model", ""))
        ws.cell(row=r_idx, column=5, value=int(r.get("quantity") or 1))
        ws.cell(row=r_idx, column=6, value=r.get("serial_number", ""))
        ws.cell(row=r_idx, column=7, value=r.get("invoice_number", ""))
        ws.cell(row=r_idx, column=8, value="Yes")
        ws.cell(row=r_idx, column=9, value=str(r.get("purchase_date", ""))[:10])
        ws.cell(row=r_idx, column=10, value=float(r.get("amount") or 0.0))
        ws.cell(row=r_idx, column=11, value=float(r.get("udhari_amount") or 0.0))
        ws.cell(row=r_idx, column=12, value=_normalize_payment_mode(r.get("payment_mode")))
        ws.cell(row=r_idx, column=13, value=str(r.get("created_at", ""))[:10])
        
    _apply_table_formatting(ws, start_row=7, headers=headers, currency_cols=[10, 11], date_cols=[9, 13], number_cols=[5])


def _build_shop_payments_sheet(ws, rows: list[dict]) -> None:
    total_amount = sum(float(r.get("total_amount") or 0.0) for r in rows)
    paid_amount = sum(float(r.get("paid_amount") or 0.0) for r in rows)
    pending_amount = sum(float(r.get("pending_amount") or 0.0) for r in rows)
    
    cards = [
        {"label": "Total Ledger Accumulation", "value": total_amount, "is_currency": True},
        {"label": "Total Udhari Paid", "value": paid_amount, "is_currency": True},
        {"label": "Outstanding Udhari Balance", "value": pending_amount, "is_currency": True}
    ]
    _write_summary_cards(ws, "SHOP PAYMENTS LEDGER", "Summarized payments and credit statements per wholesale shop", cards)
    
    headers = ["Shop Name", "Total Amount", "Paid Amount", "Pending Amount", "Settled", "Created At"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=7, column=col_idx, value=h)
        
    sorted_rows = sorted(rows, key=lambda x: x.get("created_at") or "", reverse=True)
    for r_idx, r in enumerate(sorted_rows, 8):
        shop = r.get("shops") or r.get("shop") or {}
        ws.cell(row=r_idx, column=1, value=shop.get("shop_name", ""))
        ws.cell(row=r_idx, column=2, value=float(r.get("total_amount") or 0.0))
        ws.cell(row=r_idx, column=3, value=float(r.get("paid_amount") or 0.0))
        ws.cell(row=r_idx, column=4, value=float(r.get("pending_amount") or 0.0))
        ws.cell(row=r_idx, column=5, value="Yes" if r.get("is_settled") else "No")
        ws.cell(row=r_idx, column=6, value=str(r.get("created_at", ""))[:10])
        
    _apply_table_formatting(ws, start_row=7, headers=headers, currency_cols=[2, 3, 4], date_cols=[6])


def _build_activity_logs_sheet(ws, rows: list[dict]) -> None:
    cards = [{"label": "Total Action Logs", "value": len(rows)}]
    _write_summary_cards(ws, "SYSTEM ACTIVITY LOGS", "Audit trail of user activity and system events", cards)
    
    headers = ["Action", "Device", "Created At"]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=7, column=col_idx, value=h)
        
    sorted_rows = sorted(rows, key=lambda x: x.get("created_at") or "", reverse=True)
    for r_idx, r in enumerate(sorted_rows, 8):
        ws.cell(row=r_idx, column=1, value=r.get("action", ""))
        ws.cell(row=r_idx, column=2, value=r.get("device", ""))
        ws.cell(row=r_idx, column=3, value=str(r.get("created_at", ""))[:19].replace("T", " "))
        
    _apply_table_formatting(ws, start_row=7, headers=headers, date_cols=[3])

# ---------------------------------------------------------------------------
# Public export function
# ---------------------------------------------------------------------------

def generate_excel(
    db: Client,
    export_type: str = "all",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default active sheet
    
    # We fetch all required data, applying general non-archive filters.
    # We will do date filtering inside Python for exact consistency and cross-table matches.
    
    # Let's define loaders
    def get_customers():
        return db.table("customers").select("*, batteries(*)").eq("is_archived", False).execute().data or []
        
    def get_batteries():
        return db.table("batteries").select("*, customers(*), payments(*)").eq("is_archived", False).execute().data or []
        
    def get_payments():
        return db.table("payments").select("*, customers(*), batteries(*)").eq("is_archived", False).execute().data or []
        
    def get_transactions():
        return db.table("payment_transactions").select("*, customers(*), payments(*, batteries(*))").execute().data or []
        
    def get_reminders():
        return db.table("service_reminders").select("*").eq("is_archived", False).execute().data or []
        
    def get_stock():
        return db.table("battery_stock").select("*").eq("is_archived", False).execute().data or []
        
    def get_shops():
        return db.table("shops").select("*").eq("is_archived", False).execute().data or []
        
    def get_shop_purchases():
        return db.table("shop_purchases").select("*, shops(shop_name, owner_name, mobile)").execute().data or []
        
    def get_shop_payments():
        return db.table("shop_payments").select("*, shops(shop_name)").execute().data or []
        
    def get_shop_transactions():
        return db.table("shop_payment_transactions").select("*, shops(shop_name)").execute().data or []
        
    def get_activity_logs():
        return db.table("activity_logs").select("*").execute().data or []

    # 1. Complete export "all"
    if export_type == "all":
        # Load everything
        customers = get_customers()
        batteries = get_batteries()
        payments = get_payments()
        txs = get_transactions()
        rems = get_reminders()
        stock = get_stock()
        shops = get_shops()
        sh_purchases = get_shop_purchases()
        sh_payments = get_shop_payments()
        sh_txs = get_shop_transactions()
        
        # Apply python date filters for the individual sheets in complete export if date filters exist
        f_customers = [c for c in customers if _is_in_date_range(c.get("created_at"), date_from, date_to)]
        f_batteries = [b for b in batteries if _is_in_date_range(b.get("sale_date"), date_from, date_to)]
        f_payments = [p for p in payments if _is_in_date_range(p.get("created_at"), date_from, date_to)]
        f_txs = [tx for tx in txs if _is_in_date_range(tx.get("created_at"), date_from, date_to)]
        f_rems = [r for r in rems if _is_in_date_range(r.get("reminder_date"), date_from, date_to)]
        f_stock = [s for s in stock if _is_in_date_range(s.get("created_at"), date_from, date_to)]
        f_shops = [s for s in shops if _is_in_date_range(s.get("created_at"), date_from, date_to)]
        
        f_sh_purchases = [r for r in sh_purchases if _is_in_date_range(r.get("purchase_date"), date_from, date_to)]
        f_sh_payments = [r for r in sh_payments if _is_in_date_range(r.get("created_at"), date_from, date_to)]
        f_sh_txs = [r for r in sh_txs if _is_in_date_range(r.get("created_at"), date_from, date_to)]
        
        # Generate Aggregated Customers list (from unfiltered to catch all references, then filtered)
        # Note: Aggregation is run on unfiltered list to ensure all payments/txs are linked.
        # Then, we filter the aggregated customers based on date range.
        agg_customers = _get_aggregated_customers(customers, payments, txs, rems)
        
        # Standalone scrap battery collections list
        # Scrap batteries is the subset of customers where scrap_received_date is not null
        scrap_customers = [c for c in agg_customers if c.get("scrap_received_date") and c["scrap_received_date"] != "N/A"]
        
        # Filter scrap by date if requested
        if date_from or date_to:
            scrap_customers = [c for c in scrap_customers if _is_in_date_range(c.get("scrap_received_date"), date_from, date_to)]
            agg_customers_filtered = [
                c for c in agg_customers 
                if _is_in_date_range(c["created_at"], date_from, date_to) or _is_in_date_range(c["last_activity_date"], date_from, date_to)
            ]
        else:
            agg_customers_filtered = agg_customers
            
        # Sheet 1: Dashboard Summary
        ws_dashboard = wb.create_sheet("SHOP BUSINESS SUMMARY")
        _build_business_summary_sheet(ws_dashboard, customers, batteries, payments, scrap_customers, shops, stock, f_txs)
        
        # Sheet 2: Customers
        ws_cust = wb.create_sheet("Customers")
        _build_customers_sheet(ws_cust, agg_customers_filtered)
        
        # Sheet 3: Customer Payment History (Summary & Detail)
        ws_summary = wb.create_sheet("Customer Summary")
        _build_customer_summary_sheet(ws_summary, agg_customers_filtered)
        
        ws_tx = wb.create_sheet("Transaction History")
        _build_customer_transactions_sheet(ws_tx, f_txs)
        
        # Sheet 4: Udhari
        ws_udhari = wb.create_sheet("Udhari")
        _build_payments_sheet(ws_udhari, agg_customers_filtered)
        
        # Sheet 5: Warranty Registry
        ws_warranty = wb.create_sheet("Warranty Registry")
        _build_batteries_sheet(ws_warranty, f_batteries)
        
        # Sheet 6: Scrap Batteries
        ws_scrap = wb.create_sheet("Scrap Batteries")
        _build_scrap_payments_sheet(ws_scrap, agg_customers)  # scrap_payments uses the aggregated scrap payouts
        
        # Sheet 7: Inventory
        ws_inv = wb.create_sheet("Inventory")
        _build_stock_sheet(ws_inv, f_stock)
        
        # Sheet 8: Shops / Retailers
        ws_shops = wb.create_sheet("Shops Retailers")
        _build_shops_sheet(ws_shops, f_shops, sh_purchases, sh_payments, sh_txs)
        
        # Sheet 9: Shop Transactions
        ws_shop_tx = wb.create_sheet("Shop Transactions")
        _build_shop_payment_transactions_sheet(ws_shop_tx, f_sh_txs)
        
    else:
        # Standard individual sheet export
        if export_type == "customers":
            customers = get_customers()
            payments = get_payments()
            txs = get_transactions()
            rems = get_reminders()
            agg_customers = _get_aggregated_customers(customers, payments, txs, rems)
            
            if date_from or date_to:
                agg_customers = [
                    c for c in agg_customers 
                    if _is_in_date_range(c["created_at"], date_from, date_to) or _is_in_date_range(c["last_activity_date"], date_from, date_to)
                ]
            ws = wb.create_sheet("Customers")
            _build_customers_sheet(ws, agg_customers)
            
        elif export_type == "batteries":
            batteries = get_batteries()
            if date_from or date_to:
                batteries = [b for b in batteries if _is_in_date_range(b.get("sale_date"), date_from, date_to)]
            ws = wb.create_sheet("Warranty Registry")
            _build_batteries_sheet(ws, batteries)
            
        elif export_type == "payments":
            customers = get_customers()
            payments = get_payments()
            txs = get_transactions()
            rems = get_reminders()
            agg_customers = _get_aggregated_customers(customers, payments, txs, rems)
            
            if date_from or date_to:
                agg_customers = [
                    c for c in agg_customers
                    if any(_is_in_date_range(p.get("created_at"), date_from, date_to) for p in c["payments"])
                ]
            ws = wb.create_sheet("Udhari")
            _build_payments_sheet(ws, agg_customers)
            
        elif export_type == "stock":
            stock = get_stock()
            if date_from or date_to:
                stock = [s for s in stock if _is_in_date_range(s.get("created_at"), date_from, date_to)]
            ws = wb.create_sheet("Inventory")
            _build_stock_sheet(ws, stock)
            
        elif export_type == "reminders":
            rems = get_reminders()
            if date_from or date_to:
                rems = [r for r in rems if _is_in_date_range(r.get("reminder_date"), date_from, date_to)]
            ws = wb.create_sheet("Reminders")
            _build_reminders_sheet(ws, rems)
            
        elif export_type == "shops":
            shops = get_shops()
            sh_purchases = get_shop_purchases()
            sh_payments = get_shop_payments()
            sh_txs = get_shop_transactions()
            
            if date_from or date_to:
                filtered_shops = []
                for s in shops:
                    sid_str = str(s["id"])
                    s_pur = [p for p in sh_purchases if str(p.get("shop_id")) == sid_str]
                    s_t = [t for t in sh_txs if str(t.get("shop_id")) == sid_str]
                    has_activity = (
                        any(_is_in_date_range(p.get("purchase_date"), date_from, date_to) for p in s_pur) or
                        any(_is_in_date_range(t.get("created_at"), date_from, date_to) for t in s_t) or
                        _is_in_date_range(s.get("created_at"), date_from, date_to)
                    )
                    if has_activity:
                        filtered_shops.append(s)
                shops = filtered_shops
            ws = wb.create_sheet("Shops")
            _build_shops_sheet(ws, shops, sh_purchases, sh_payments, sh_txs)
            
        elif export_type == "shop_purchases":
            sh_purchases = get_shop_purchases()
            if date_from or date_to:
                sh_purchases = [r for r in sh_purchases if _is_in_date_range(r.get("purchase_date"), date_from, date_to)]
            ws = wb.create_sheet("Shop Purchases")
            _build_shop_purchases_sheet(ws, sh_purchases)
            
        elif export_type == "shop_payments":
            sh_payments = get_shop_payments()
            if date_from or date_to:
                sh_payments = [r for r in sh_payments if _is_in_date_range(r.get("created_at"), date_from, date_to)]
            ws = wb.create_sheet("Shop Payments")
            _build_shop_payments_sheet(ws, sh_payments)
            
        elif export_type == "shop_payment_transactions":
            sh_txs = get_shop_transactions()
            if date_from or date_to:
                sh_txs = [r for r in sh_txs if _is_in_date_range(r.get("created_at"), date_from, date_to)]
            ws = wb.create_sheet("Shop Transactions")
            _build_shop_payment_transactions_sheet(ws, sh_txs)
            
        elif export_type == "activity_logs":
            logs = get_activity_logs()
            if date_from or date_to:
                logs = [r for r in logs if _is_in_date_range(r.get("created_at"), date_from, date_to)]
            ws = wb.create_sheet("Activity Logs")
            _build_activity_logs_sheet(ws, logs)
            
        elif export_type == "customer_payment_transactions":
            # Generates Customer Summary & Transaction History tabs
            customers = get_customers()
            payments = get_payments()
            txs = get_transactions()
            rems = get_reminders()
            
            agg_customers = _get_aggregated_customers(customers, payments, txs, rems)
            
            if date_from or date_to:
                agg_customers = [
                    c for c in agg_customers 
                    if _is_in_date_range(c["created_at"], date_from, date_to) or _is_in_date_range(c["last_activity_date"], date_from, date_to)
                ]
                txs = [tx for tx in txs if _is_in_date_range(tx.get("created_at"), date_from, date_to)]
                
            ws_summary = wb.create_sheet("Customer Summary")
            _build_customer_summary_sheet(ws_summary, agg_customers)
            
            ws_tx = wb.create_sheet("Transaction History")
            _build_customer_transactions_sheet(ws_tx, txs)
            
        elif export_type == "scrap_payments":
            customers = get_customers()
            payments = get_payments()
            txs = get_transactions()
            rems = get_reminders()
            agg_customers = _get_aggregated_customers(customers, payments, txs, rems)
            
            if date_from or date_to:
                agg_customers = [
                    c for c in agg_customers 
                    if _is_in_date_range(c.get("scrap_received_date"), date_from, date_to)
                ]
            ws = wb.create_sheet("Scrap Payments")
            _build_scrap_payments_sheet(ws, agg_customers)
            
        else:
            raise ValueError(f"Unknown export_type: '{export_type}'")
            
    if not wb.sheetnames:
        raise ValueError(f"Unknown export_type: '{export_type}'")
        
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_shop_statement_excel(db: Client, shop_id: str) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet
    
    from app.services.shop import get_shop_details
    from app.database import safe_execute
    details = get_shop_details(db, shop_id)
    shop = details["shop"]
    purchases = details["purchases"]
    payment = details["payment"]
    transactions = details["transactions"]
    
    # Sort history records latest first
    purchases = sorted(purchases, key=lambda x: getattr(x, "purchase_date", "") or "", reverse=True)
    transactions = sorted(transactions, key=lambda x: getattr(x, "created_at", "") or "", reverse=True)
    
    # 1. Statement Summary
    ws_summary = wb.create_sheet("Statement Summary")
    ws_summary.sheet_view.showGridLines = True
    
    ws_summary.append(["SHREE GANADHISH BATTERY SERVICES"])
    ws_summary.append(["SHOP STATEMENT OF ACCOUNT"])
    ws_summary.append([])
    
    ws_summary.append(["Shop Name:", shop.shop_name])
    ws_summary.append(["Owner Name:", shop.owner_name])
    ws_summary.append(["Mobile Number:", shop.mobile])
    ws_summary.append(["Address:", shop.address or "N/A"])
    ws_summary.append([])
    
    total_pur = sum(float(p.amount) for p in purchases)
    total_udhari = float(payment.total_amount) if payment else 0.0
    paid_udhari = float(payment.paid_amount) if payment else 0.0
    pending_udhari = float(payment.pending_amount) if payment else 0.0
    
    ws_summary.append(["FINANCIAL SUMMARY"])
    ws_summary.append(["Total Purchase Value:", total_pur])
    ws_summary.append(["Total Udhari Accumulation:", total_udhari])
    ws_summary.append(["Total Udhari Paid:", paid_udhari])
    ws_summary.append(["Outstanding Udhari Balance:", pending_udhari])
    
    ws_summary["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws_summary["A2"].font = Font(bold=True, size=12)
    ws_summary["A9"].font = Font(bold=True, size=12, color="1E3A5F")
    for row in range(10, 14):
        ws_summary[f"A{row}"].font = Font(bold=True)
        cell = ws_summary[f"B{row}"]
        cell.number_format = '"₹"#,##0.00'
        cell.alignment = Alignment(horizontal="right")
        
    _apply_table_formatting(ws_summary, start_row=9, headers=["FINANCIAL SUMMARY", "VALUES"], currency_cols=[2])
    
    # 2. Purchases History
    ws_purchases = wb.create_sheet("Purchase History")
    p_headers = ["Purchase Date", "Battery Model", "Quantity", "Serial Numbers", "Invoice Number", "Stock Reduced", "Amount (₹)", "Udhari Amount (₹)", "Payment Mode"]
    for col_idx, h in enumerate(p_headers, 1):
        ws_purchases.cell(row=1, column=col_idx, value=h)
        
    for p in purchases:
        ws_purchases.append([
            str(p.purchase_date)[:10],
            p.battery_model,
            p.quantity,
            p.serial_number,
            p.invoice_number,
            "Yes",
            float(p.amount),
            float(p.udhari_amount),
            _normalize_payment_mode(p.payment_mode)
        ])
    _apply_table_formatting(ws_purchases, start_row=1, headers=p_headers, currency_cols=[7, 8], date_cols=[1], number_cols=[3])
    
    # 3. Udhari Ledger Transaction History
    ws_payments = wb.create_sheet("Udhari Transaction History")
    t_headers = ["Transaction Date", "Transaction Type", "Amount (₹)", "Notes", "Payment Mode", "Payment Date"]
    for col_idx, h in enumerate(t_headers, 1):
        ws_payments.cell(row=1, column=col_idx, value=h)
        
    for t in transactions:
        ws_payments.append([
            str(t.created_at)[:10],
            t.transaction_type,
            float(t.amount),
            t.notes or "",
            _normalize_payment_mode(t.payment_mode),
            str(t.created_at)[:10],
        ])
    _apply_table_formatting(ws_payments, start_row=1, headers=t_headers, currency_cols=[3], date_cols=[1, 6])
    
    # 4. Stock Movement
    ws_movement = wb.create_sheet("Stock Movement")
    m_headers = ["Battery Model", "Opening Stock", "Quantity Sold (To Shop)", "Closing Stock"]
    for col_idx, h in enumerate(m_headers, 1):
        ws_movement.cell(row=1, column=col_idx, value=h)
        
    from collections import defaultdict
    model_sold = defaultdict(int)
    for p in purchases:
        model_sold[p.battery_model.strip().upper()] += int(p.quantity or 1)
        
    stock_res = safe_execute(db.table("battery_stock").select("model_name, quantity").eq("is_archived", False))
    stock_map = {s["model_name"].strip().upper(): s["quantity"] for s in (stock_res.data or [])}
    
    for model, qty_sold in sorted(model_sold.items()):
        closing = stock_map.get(model, 0)
        opening = closing + qty_sold
        ws_movement.append([
            model,
            opening,
            qty_sold,
            closing
        ])
    _apply_table_formatting(ws_movement, start_row=1, headers=m_headers, number_cols=[2, 3, 4])
    
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def archive_old_records(db: Client, year: int, action: str) -> dict:
    """
    Implements Day 9 & Day 10 Smart Cleanup Logic:
    - action: 'archive' (set is_archived = True) or 'delete' (permanently delete)
    - year: target year (archives/deletes data from that year or before)
    
    Protections:
    - DO NOT archive/delete active warranties (warranty_expiry >= today).
    - DO NOT archive/delete pending udhari (payments where is_settled = False).
    - DO NOT archive/delete customers who have pending udhari or active warranties.
    - DO NOT archive/delete reminders if:
        * warranty is still active
        * future reminder exists
        * reminder status is pending/uncompleted
        * customer has active service cycle
    """
    today = datetime.date.today().isoformat()
    ninety_days_ago = (datetime.date.today() - datetime.timedelta(days=90)).isoformat()
    six_months_ago = (datetime.date.today() - datetime.timedelta(days=180)).isoformat()
    
    # Automatically prune activity logs older than 90 days (LIGHTWEIGHT ACTIVITY FEED POLICY)
    try:
        db.table("activity_logs").delete().lt("created_at", f"{ninety_days_ago}T00:00:00+00:00").execute()
    except Exception:
        pass

    # Prune activity logs to retain only important business events
    try:
        logs_res = db.table("activity_logs").select("id, action").execute()
        important_prefixes = (
            "BATTERY_ADDED", "PAYMENT_ADDED", "STOCK_ADJUSTED",
            "STOCK_INCREASED", "STOCK_DECREASED", "STOCK_AUTO_DECREASED",
            "REMINDER_COMPLETED", "EMAIL_BACKUP_SENT", "EMAIL_BACKUP_FAILED",
            "BATTERY_RETURNED", "SHOP_PURCHASE_DELETED", "OPENING_BALANCE_ADDED",
            "ADJUSTMENT_ADDED"
        )
        to_delete = []
        for log in (logs_res.data or []):
            action_str = log.get("action", "") or ""
            if not any(action_str.startswith(pref) for pref in important_prefixes):
                to_delete.append(log["id"])
        
        if to_delete:
            for i in range(0, len(to_delete), 100):
                chunk = to_delete[i:i+100]
                db.table("activity_logs").delete().in_("id", chunk).execute()
    except Exception:
        pass

    # Auto-delete completed reminders older than 90 days to prevent table growth
    try:
        ninety_days_ago = (datetime.date.today() - datetime.timedelta(days=90)).isoformat()
        db.table("service_reminders").delete()\
            .eq("is_completed", True)\
            .lt("reminder_date", ninety_days_ago)\
            .execute()
    except Exception:
        pass

    
    # 1. Fetch protected customer IDs (due to active warranties or pending payments)
    # A. Customers with active warranties
    active_b_res = db.table("batteries").select("customer_id").gte("warranty_expiry", today).eq("is_archived", False).execute()
    active_b_cust_ids = {str(row["customer_id"]) for row in (active_b_res.data or [])}
    
    # B. Customers with pending payments
    pending_p_res = db.table("payments").select("customer_id").eq("is_settled", False).eq("is_archived", False).execute()
    pending_p_cust_ids = {str(row["customer_id"]) for row in (pending_p_res.data or [])}
    
    # C. Reminders check for active service cycles / future / pending reminders
    rem_status_res = db.table("service_reminders")\
        .select("battery_id, customer_id, is_completed, reminder_date, warranty_expiry")\
        .execute()
    
    protected_batteries_rems = set()
    protected_customers_rems = set()
    for row in (rem_status_res.data or []):
        bid = str(row.get("battery_id")) if row.get("battery_id") else None
        cid = str(row.get("customer_id")) if row.get("customer_id") else None
        is_comp = row.get("is_completed", False)
        rem_d = str(row.get("reminder_date", ""))[:10]
        w_exp = str(row.get("warranty_expiry", ""))[:10] if row.get("warranty_expiry") else ""
        
        # Check active warranty
        w_active = w_exp >= today if w_exp else False
        # Check future reminder
        rem_future = rem_d > today if rem_d else False
        # Check pending reminder
        rem_pending = not is_comp
        
        if bid and (w_active or rem_future or rem_pending):
            protected_batteries_rems.add(bid)
        if cid and (w_active or rem_future or rem_pending):
            protected_customers_rems.add(cid)

    # Combined protected customer set (warranties + payments + reminders active service cycle)
    protected_cust_ids = active_b_cust_ids.union(pending_p_cust_ids).union(protected_customers_rems)
    
    # Define date range for target year: up to the end of that year (YYYY-12-31)
    end_date = f"{year}-12-31"
    
    # Initialize count counters
    payments_processed = 0
    batteries_processed = 0
    customers_processed = 0
    reminders_processed = 0
    
    if action == "archive":
        # 1. Archive settled payments created in target year or before
        target_payments = db.table("payments")\
            .select("id")\
            .lte("created_at", f"{end_date}T23:59:59+00:00")\
            .eq("is_settled", True)\
            .eq("is_archived", False)\
            .execute()
        target_payment_ids = [str(r["id"]) for r in (target_payments.data or [])]
        if target_payment_ids:
            db.table("payments").update({"is_archived": True}).in_("id", target_payment_ids).execute()
            payments_processed = len(target_payment_ids)
            
        # 2. Archive expired batteries sold in target year or before
        target_batteries = db.table("batteries")\
            .select("id")\
            .lte("sale_date", end_date)\
            .lt("warranty_expiry", today)\
            .eq("is_archived", False)\
            .execute()
        target_battery_ids = [str(r["id"]) for r in (target_batteries.data or [])]
        if target_battery_ids:
            db.table("batteries").update({"is_archived": True}).in_("id", target_battery_ids).execute()
            batteries_processed = len(target_battery_ids)
            
        # 3. Archive inactive customers created in target year or before
        target_customers = db.table("customers")\
            .select("id")\
            .lte("created_at", f"{end_date}T23:59:59+00:00")\
            .eq("is_archived", False)\
            .execute()
        target_customer_ids = [str(r["id"]) for r in (target_customers.data or [])]
        unprotected_customer_ids = [cid for cid in target_customer_ids if cid not in protected_cust_ids]
        if unprotected_customer_ids:
            db.table("customers").update({"is_archived": True}).in_("id", unprotected_customer_ids).execute()
            customers_processed = len(unprotected_customer_ids)

        # 4. Archive eligible completed reminders falling in target year or before (older than 90 days)
        target_reminders = db.table("service_reminders")\
            .select("id, reminder_date")\
            .lte("reminder_date", end_date)\
            .eq("is_completed", True)\
            .eq("is_archived", False)\
            .execute()
        eligible_reminder_ids = []
        for r in (target_reminders.data or []):
            rid = str(r["id"])
            r_date = str(r["reminder_date"])[:10]
            if r_date < ninety_days_ago:
                eligible_reminder_ids.append(rid)
                
        if eligible_reminder_ids:
            db.table("service_reminders").update({"is_archived": True}).in_("id", eligible_reminder_ids).execute()
            reminders_processed = len(eligible_reminder_ids)
            
    elif action == "delete":
        # 1. Delete settled payments created in target year or before
        target_payments = db.table("payments")\
            .select("id")\
            .lte("created_at", f"{end_date}T23:59:59+00:00")\
            .eq("is_settled", True)\
            .execute()
        target_payment_ids = [str(r["id"]) for r in (target_payments.data or [])]
        if target_payment_ids:
            db.table("payments").delete().in_("id", target_payment_ids).execute()
            payments_processed = len(target_payment_ids)
            
        # 2. Delete expired batteries sold in target year or before
        target_batteries = db.table("batteries")\
            .select("id")\
            .lte("sale_date", end_date)\
            .lt("warranty_expiry", today)\
            .execute()
        target_battery_ids = [str(r["id"]) for r in (target_batteries.data or [])]
        if target_battery_ids:
            ref_payments = db.table("payments").select("battery_id").in_("battery_id", target_battery_ids).execute()
            ref_battery_ids = {str(row["battery_id"]) for row in (ref_payments.data or []) if row.get("battery_id")}
            safe_battery_ids = [bid for bid in target_battery_ids if bid not in ref_battery_ids]
            if safe_battery_ids:
                db.table("batteries").delete().in_("id", safe_battery_ids).execute()
                batteries_processed = len(safe_battery_ids)
                
        # 3. Delete inactive customers created in target year or before
        target_customers = db.table("customers")\
            .select("id")\
            .lte("created_at", f"{end_date}T23:59:59+00:00")\
            .execute()
        target_customer_ids = [str(r["id"]) for r in (target_customers.data or [])]
        unprotected_customer_ids = [cid for cid in target_customer_ids if cid not in protected_cust_ids]
        if unprotected_customer_ids:
            ref_batteries = db.table("batteries").select("customer_id").in_("customer_id", unprotected_customer_ids).execute()
            ref_cust_ids = {str(row["customer_id"]) for row in (ref_batteries.data or [])}
            safe_cust_ids = [cid for cid in unprotected_customer_ids if cid not in ref_cust_ids]
            if safe_cust_ids:
                db.table("customers").delete().in_("id", safe_cust_ids).execute()
                customers_processed = len(safe_cust_ids)

        # 4. Delete eligible completed reminders falling in target year or before (older than 90 days)
        target_reminders = db.table("service_reminders")\
            .select("id, reminder_date")\
            .lte("reminder_date", end_date)\
            .eq("is_completed", True)\
            .execute()
        eligible_reminder_ids = []
        for r in (target_reminders.data or []):
            rid = str(r["id"])
            r_date = str(r["reminder_date"])[:10]
            if r_date < ninety_days_ago:
                eligible_reminder_ids.append(rid)
                
        if eligible_reminder_ids:
            db.table("service_reminders").delete().in_("id", eligible_reminder_ids).execute()
            reminders_processed = len(eligible_reminder_ids)

    return {
        "payments_count": payments_processed,
        "batteries_count": batteries_processed,
        "customers_count": customers_processed,
        "reminders_count": reminders_processed,
    }
